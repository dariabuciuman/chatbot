import time

import openpyxl


def clean_up_diacritics(sentence):
    special_chars = {
        'Äƒ': 'a',
        'Ã¢': 'a',
        'Ã®': 'i',
        'È™': 's',
        'È›': 't',
        'Ä‚': 'A',
        'Ã‚': 'A',
        'ÃŽ': 'I',
        'È˜': 'S',
        'Èš': 'T'
    }
    for special_char, regular_char in special_chars.items():
        sentence = sentence.replace(special_char, regular_char)
    return sentence


def generate_excel(text):
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # Set column labels
    sheet.cell(row=1, column=1).value = "Text"
    sheet.cell(row=1, column=2).value = "Category"

    # rows = input.split("\n")
    for i, row in enumerate(text):
        temp_line = clean_up_diacritics(row)
        parts = temp_line.split('"')
        print(parts)
        if len(parts) == 3:
            first_column_text = parts[1].strip()
            second_column_text = parts[0].strip() + parts[2].strip()
            sheet.cell(row=i + 2, column=1).value = first_column_text
            sheet.cell(row=i + 2, column=2).value = second_column_text
        else:
            sheet.cell(row=i + 2, column=1).value = row.strip()

    # Remove unused columns
    sheet.delete_cols(3, sheet.max_column - 2)

    # Save the workbook with a timestamp in the filename
    timestamp = int(time.time())
    workbook.save("output_" + str(timestamp) + ".xlsx")
    print("Excel file generated successfully!")


file_path = "C:/Users/buciu/Desktop/penal-code/dataset.txt"  # Replace with the actual file path

with open(file_path, 'r') as file:
    lines = file.readlines()
    generate_excel(lines)
